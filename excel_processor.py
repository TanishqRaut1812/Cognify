import openpyxl
from database import get_db
from models import recalculate_scores_and_rankings

def validate_excel_results(file_stream, test_id):
    """
    Parses and validates an uploaded .xlsx file for a test.
    Checks against master student database and test total marks.
    Returns preview dict with error list and data ready for publishing.
    """
    conn = get_db()
    cursor = conn.cursor()

    # Get test info
    test = cursor.execute('SELECT * FROM tests WHERE id = ?', (test_id,)).fetchone()
    if not test:
        conn.close()
        return {'valid': False, 'errors': ['Test not found.']}

    total_marks = float(test['total_marks'])

    # Pre-fetch all master students
    master_students = cursor.execute('SELECT registration_no, roll_no, name, class_name FROM students').fetchall()
    master_reg_map = {s['registration_no'].strip().upper(): s for s in master_students}

    conn.close()

    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = wb.active
    except Exception as e:
        return {'valid': False, 'errors': [f'Failed to read Excel file: {str(e)}']}

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {'valid': False, 'errors': ['The uploaded Excel sheet is empty.']}

    # Header index mapping
    header_row = None
    header_idx = -1
    for idx, row in enumerate(rows):
        if not row:
            continue
        row_str = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        if any('registration' in cell or 'reg' in cell for cell in row_str):
            header_row = row_str
            header_idx = idx
            break

    if header_row is None:
        return {'valid': False, 'errors': ['Could not locate header row. Expected columns: Registration Number, Roll Number, Name, Attendance, Score.']}

    col_reg = -1
    col_roll = -1
    col_name = -1
    col_att = -1
    col_score = -1

    for c_idx, cell in enumerate(header_row):
        if 'reg' in cell:
            col_reg = c_idx
        elif 'roll' in cell:
            col_roll = c_idx
        elif 'name' in cell:
            col_name = c_idx
        elif 'attend' in cell or 'status' in cell:
            col_att = c_idx
        elif 'score' in cell or 'mark' in cell or 'pct' in cell:
            col_score = c_idx

    if col_reg == -1 or col_att == -1 or col_score == -1:
        return {
            'valid': False,
            'errors': ['Missing required columns in Excel header. Expected: Registration Number, Attendance, Score (Marks Obtained).']
        }

    errors = []
    warnings = []
    seen_regs = set()
    duplicate_records = 0
    invalid_scores = 0
    present_count = 0
    absent_count = 0

    parsed_records = []

    for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        raw_reg = str(row[col_reg]).strip() if col_reg < len(row) and row[col_reg] is not None else ''
        if not raw_reg or raw_reg.lower() in ('none', 'null'):
            continue

        reg_norm = raw_reg.upper()

        # Check duplicate in sheet
        if reg_norm in seen_regs:
            errors.append(f'Row {row_no}: Duplicate student registration number "{raw_reg}" in sheet.')
            duplicate_records += 1
            continue
        seen_regs.add(reg_norm)

        # Check against master student database
        if reg_norm not in master_reg_map:
            errors.append(f'Row {row_no}: Registration number "{raw_reg}" not found in Master Student Database.')
            continue

        master_student = master_reg_map[reg_norm]

        # Attendance parsing
        raw_att = str(row[col_att]).strip() if col_att < len(row) and row[col_att] is not None else ''
        att_lower = raw_att.lower()

        if att_lower in ('present', 'p', '1', 'yes', 'true'):
            attendance = 'Present'
            present_count += 1
        elif att_lower in ('absent', 'a', '0', 'no', 'false'):
            attendance = 'Absent'
            absent_count += 1
        else:
            errors.append(f'Row {row_no}: Invalid attendance value "{raw_att}". Expected "Present" or "Absent".')
            continue

        # Score parsing
        raw_score = row[col_score] if col_score < len(row) else 0
        if attendance == 'Absent':
            marks_obtained = 0.0
            percentage = 0.0
        else:
            try:
                marks_obtained = float(raw_score)
            except (ValueError, TypeError):
                errors.append(f'Row {row_no}: Invalid numeric score "{raw_score}" for student {master_student["name"]}.')
                invalid_scores += 1
                continue

            if marks_obtained < 0:
                errors.append(f'Row {row_no}: Score cannot be negative ({marks_obtained}) for student {master_student["name"]}.')
                invalid_scores += 1
                continue

            if marks_obtained > total_marks:
                errors.append(f'Row {row_no}: Score {marks_obtained} exceeds test total marks ({total_marks}) for student {master_student["name"]}.')
                invalid_scores += 1
                continue

            percentage = round((marks_obtained / total_marks) * 100.0, 2)

        parsed_records.append({
            'registration_no': master_student['registration_no'],
            'attendance': attendance,
            'marks_obtained': marks_obtained,
            'percentage': percentage
        })

    # Check for missing students from master DB
    missing_regs = set(master_reg_map.keys()) - seen_regs
    if missing_regs:
        warnings.append(f'{len(missing_regs)} master students were not listed in this sheet and will be marked Absent (0%).')

    total_detected = len(seen_regs)
    total_master = len(master_reg_map)

    is_valid = len(errors) == 0 and len(parsed_records) > 0

    return {
        'valid': is_valid,
        'test_id': test_id,
        'test_name': test['test_name'],
        'total_detected': total_detected,
        'total_master': total_master,
        'present_count': present_count,
        'absent_count': absent_count + len(missing_regs),
        'valid_records': len(parsed_records),
        'duplicate_records': duplicate_records,
        'invalid_scores': invalid_scores,
        'errors': errors,
        'warnings': warnings,
        'parsed_records': parsed_records,
        'missing_regs': list(missing_regs)
    }

def publish_test_results(test_id, parsed_records, missing_regs=None):
    """
    Atomically saves/replaces test results in the database and triggers score recalculation.
    """
    conn = get_db()
    cursor = conn.cursor()

    try:
        # Atomic transaction
        cursor.execute('BEGIN TRANSACTION;')

        # Clear existing results for this test
        cursor.execute('DELETE FROM test_results WHERE test_id = ?', (test_id,))

        # Insert parsed records
        for r in parsed_records:
            cursor.execute('''
                INSERT INTO test_results (test_id, registration_no, attendance, marks_obtained, percentage)
                VALUES (?, ?, ?, ?, ?)
            ''', (test_id, r['registration_no'], r['attendance'], r['marks_obtained'], r['percentage']))

        # Auto-insert missing students as Absent (0 marks)
        if missing_regs:
            for reg in missing_regs:
                cursor.execute('''
                    INSERT INTO test_results (test_id, registration_no, attendance, marks_obtained, percentage)
                    VALUES (?, ?, 'Absent', 0.0, 0.0)
                ''', (test_id, reg))

        # Mark test as published
        cursor.execute('UPDATE tests SET is_published = 1 WHERE id = ?', (test_id,))

        conn.commit()
        conn.close()

        # Recalculate official scores & competition rankings
        recalculate_scores_and_rankings()
        return {'success': True, 'message': 'Test results published and rankings updated successfully.'}

    except Exception as e:
        conn.rollback()
        conn.close()
        return {'success': False, 'error': f'Database transaction failed: {str(e)}'}

def validate_student_list_excel(file_stream, target_class):
    """
    Parses and validates an uploaded .xlsx file containing master students for a specific class (SY, TY, Final Year).
    Checks required columns, missing values, duplicate reg numbers, and duplicate roll numbers.
    """
    if target_class not in ('SY', 'TY', 'Final Year'):
        return {'valid': False, 'errors': [f'Invalid target class: {target_class}']}

    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = wb.active
    except Exception as e:
        return {'valid': False, 'errors': [f'Failed to read Excel file: {str(e)}']}

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {'valid': False, 'errors': ['The uploaded Excel sheet is empty.']}

    # Locate header row
    header_row = None
    header_idx = -1
    for idx, row in enumerate(rows):
        if not row:
            continue
        row_str = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        if any('registration' in cell or 'reg' in cell for cell in row_str):
            header_row = row_str
            header_idx = idx
            break

    if header_row is None:
        return {'valid': False, 'errors': ['Could not locate header row. Expected columns: Registration Number, Roll Number, Name.']}

    col_reg = -1
    col_roll = -1
    col_name = -1

    for c_idx, cell in enumerate(header_row):
        if 'reg' in cell:
            col_reg = c_idx
        elif 'roll' in cell:
            col_roll = c_idx
        elif 'name' in cell:
            col_name = c_idx

    if col_reg == -1 or col_roll == -1 or col_name == -1:
        return {
            'valid': False,
            'errors': ['Missing required header columns. Expected: Registration Number, Roll Number, Name.']
        }

    errors = []
    warnings = []
    seen_regs = set()
    seen_rolls = set()
    duplicate_reg_count = 0
    duplicate_roll_count = 0

    parsed_students = []

    for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        raw_reg = str(row[col_reg]).strip() if col_reg < len(row) and row[col_reg] is not None else ''
        raw_roll = str(row[col_roll]).strip() if col_roll < len(row) and row[col_roll] is not None else ''
        raw_name = str(row[col_name]).strip() if col_name < len(row) and row[col_name] is not None else ''

        if not raw_reg or raw_reg.lower() in ('none', 'null'):
            errors.append(f'Row {row_no}: Missing Registration Number.')
            continue
        if not raw_roll or raw_roll.lower() in ('none', 'null'):
            errors.append(f'Row {row_no}: Missing Roll Number.')
            continue
        if not raw_name or raw_name.lower() in ('none', 'null'):
            errors.append(f'Row {row_no}: Missing Name.')
            continue

        reg_norm = raw_reg.upper()
        roll_norm = raw_roll.upper()

        if reg_norm in seen_regs:
            errors.append(f'Row {row_no}: Duplicate Registration Number "{raw_reg}" in sheet.')
            duplicate_reg_count += 1
            continue
        seen_regs.add(reg_norm)

        if roll_norm in seen_rolls:
            errors.append(f'Row {row_no}: Duplicate Roll Number "{raw_roll}" in sheet.')
            duplicate_roll_count += 1
            continue
        seen_rolls.add(roll_norm)

        parsed_students.append({
            'registration_no': raw_reg,
            'roll_no': raw_roll,
            'name': raw_name,
            'class_name': target_class
        })

    is_valid = len(errors) == 0 and len(parsed_students) > 0

    return {
        'valid': is_valid,
        'class_name': target_class,
        'total_detected': len(parsed_students) + duplicate_reg_count + duplicate_roll_count + len([e for e in errors if 'Missing' in e]),
        'valid_count': len(parsed_students),
        'duplicate_reg_count': duplicate_reg_count,
        'duplicate_roll_count': duplicate_roll_count,
        'errors': errors,
        'warnings': warnings,
        'parsed_students': parsed_students
    }

def save_master_student_list(target_class, parsed_students):
    """
    Atomically replaces the master student list for target_class.
    """
    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute('BEGIN TRANSACTION;')

        # Delete existing students for this class
        cursor.execute('DELETE FROM students WHERE class_name = ?', (target_class,))

        for s in parsed_students:
            cursor.execute('''
                INSERT INTO students (registration_no, roll_no, name, class_name)
                VALUES (?, ?, ?, ?)
            ''', (s['registration_no'], s['roll_no'], s['name'], target_class))

        conn.commit()
        conn.close()

        # Recalculate rankings and Cognify scores
        recalculate_scores_and_rankings()
        return {'success': True, 'message': f'{target_class} Master Student List imported successfully ({len(parsed_students)} students).'}

    except Exception as e:
        conn.rollback()
        conn.close()
        return {'success': False, 'error': f'Failed to import student list: {str(e)}'}

